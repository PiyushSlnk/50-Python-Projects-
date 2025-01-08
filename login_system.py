import openpyxl
from openpyxl import Workbook
import os

# File path for Excel data
FILE_PATH = "user_data.xlsx"

# Initialize Excel sheet if not exists
def initialize_excel():
    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Users"
        sheet.append(["Username", "Password"])
        sheet.append(["demo_user", "demo_pass"])  # Add demo user
        wb.save(FILE_PATH)

# Load Excel sheet
def load_excel():
    if not os.path.exists(FILE_PATH):
        initialize_excel()
    return openpyxl.load_workbook(FILE_PATH)

# Function to verify login
def login(username, password):
    wb = load_excel()
    sheet = wb["Users"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            return True
    return False

# Function to reset password
def reset_password(username, new_password):
    wb = load_excel()
    sheet = wb["Users"]
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == username:
            row[1].value = new_password
            wb.save(FILE_PATH)
            return True
    return False

# Main menu
def main_menu():
    initialize_excel()
    print("Welcome to the Login System")
    while True:
        print("\nOptions:")
        print("1. Login")
        print("2. Reset Password")
        print("3. Exit")
        choice = input("Choose an option: ")

        if choice == "1":
            username = input("Enter your username: ")
            password = input("Enter your password: ")
            if login(username, password):
                print(f"Welcome, {username}! Login successful.")
            else:
                print("Invalid username or password. Please try again.")

        elif choice == "2":
            username = input("Enter your username: ")
            new_password = input("Enter your new password: ")
            if reset_password(username, new_password):
                print("Password reset successful. You can now log in with the new password.")
            else:
                print("Username not found. Please try again.")

        elif choice == "3":
            print("Exiting the system. Goodbye!")
            break

        else:
            print("Invalid option. Please choose again.")

# Run the program
if __name__ == "__main__":
    main_menu()
